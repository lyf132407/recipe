from openpyxl import Workbook
import os
from tools.recipe_crud import RecipeCRUD

def export_all_recipe(file_name="菜谱收藏.xlsx"):
    save_path = file_name
    data = RecipeCRUD._read_all()
    wb = Workbook()
    sheet = wb.active
    sheet.title = "菜谱清单"
    # 表头
    sheet["A1"] = "分类"
    sheet["B1"] = "菜名"
    sheet["C1"] = "食材"
    sheet["D1"] = "详细步骤"
    # 写入数据
    for row_idx, item in enumerate(data, start=2):
        sheet.cell(row=row_idx, column=1, value=item["分类"])
        sheet.cell(row=row_idx, column=2, value=item["菜名"])
        sheet.cell(row=row_idx, column=3, value="、".join(item["食材"]))
        sheet.cell(row=row_idx, column=4, value=item["详细步骤"])
    wb.save(save_path)
    return f"导出成功，文件路径：{os.path.abspath(save_path)}"

if __name__ == "__main__":
    print(export_all_recipe())
